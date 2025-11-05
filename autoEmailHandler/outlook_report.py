import win32com.client as win32
from datetime import datetime, timedelta
import re
from collections import OrderedDict, defaultdict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ---------------------------------------------
# 1) 连接 Outlook 并筛选最近7天未读
# ---------------------------------------------
outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")
inbox = outlook.GetDefaultFolder(6)  # 6 = olFolderInbox
items = inbox.Items
items.Sort("[ReceivedTime]", True)

since = (datetime.now() - timedelta(days=7)).strftime("%m/%d/%Y %I:%M %p")  # MM/DD/YYYY 12小时制
restriction = f"[Unread] = True AND [ReceivedTime] >= '{since}'"
filtered = items.Restrict(restriction)

print(f"未读近7天: {len(filtered)} 封")

# ---------------------------------------------
# 2) 解析函数：把正文里的 1a/1b/2a/2b ... 抓出来
#   - 兼容: 冒号(:/：)、点(.)、右括号()) 等常见分隔
#   - 关键点：用“下一段开始”为边界提取多行
# ---------------------------------------------
# 键名如：'1a', '1b', '2a', '2b' ...（不区分大小写）
KEY_TOKEN = r"(?P<key>\d{1,2}[a-dA-D])"
# 分隔符常见几种：冒号/全角冒号/点/右括号/空格
SEP_TOKEN = r"(?:\s*[:：\.\)\-]\s*|\s+)"
# 下一个键的开始，用于界定当前值的结束
NEXT_KEY_START = r"(?=\n\s*\d{1,2}[a-dA-D]\s*(?:[:：\.\)\-]|\s))"

# 用 DOTALL 让 . 能跨行
FIELD_RE = re.compile(KEY_TOKEN + SEP_TOKEN + r"(?P<val>.+?)(?:\n\s*" + NEXT_KEY_START + r"|$)", re.IGNORECASE | re.DOTALL)

def parse_numbered_fields(text: str) -> OrderedDict:
    """
    从正文里提取 1a/1b/2a/2b 等字段，返回有序字典（按出现顺序）。
    会做简单清洗：去除多余空白、换行压缩。
    """
    if not text:
        return OrderedDict()
    # 统一换行符
    body = text.replace("\r\n", "\n").replace("\r", "\n")
    results = OrderedDict()
    for m in FIELD_RE.finditer(body):
        key = m.group("key").lower()  # 统一小写：1a/1b/…
        val = m.group("val").strip()
        # 清洗：多行压成单行，但保留分隔空格
        val = re.sub(r"[ \t]*\n[ \t]*", " ", val).strip()
        results[key] = val
    return results

# ---------------------------------------------
# 3) 收集所有邮件：原始列 + 解析列
#    - raw_rows: 原始信息（Raw 工作表）
#    - parsed_rows: 解析后的行（Parsed 工作表）
#    - 动态统计所有出现过的键，保证列一致
# ---------------------------------------------
raw_rows = []
parsed_rows = []
all_keys = []                      # 保留按“首次出现顺序”的键
seen_keys = set()

for mail in filtered:
    if mail.Class != 43:  # 43 = olMail，跳过日程/回执等
        continue

    subject = mail.Subject or ""
    sender = getattr(mail, "SenderEmailAddress", "") or ""
    recv_at = mail.ReceivedTime
    # 尽量拿文本，如果需要也可用 HTMLBody
    body = (mail.Body or "").strip()

    # 解析 1a/1b/2a/2b...
    fields = parse_numbered_fields(body)

    # 记录所有见过的键，以便导出为列
    for k in fields.keys():
        if k not in seen_keys:
            seen_keys.add(k)
            all_keys.append(k)

    # Raw 工作表的行
    raw_rows.append({
        "subject": subject,
        "from": sender,
        "received": recv_at.strftime("%Y-%m-%d %H:%M:%S"),
        "preview": body[:300].replace("\n", " ").replace("\r", " ")
    })

    # Parsed 工作表的行：基础列 + 动态字段列
    row = {
        "_subject": subject,
        "_from": sender,
        "_received": recv_at.strftime("%Y-%m-%d %H:%M:%S"),
    }
    row.update(fields)  # 键如 "1a","1b","2a"...
    parsed_rows.append(row)

print(f"可解析出字段列: {all_keys}")

# ---------------------------------------------
# 4) 导出为 xlsx（Raw / Parsed 两张表）
# ---------------------------------------------
wb = Workbook()

# Sheet 1: Raw
ws_raw = wb.active
ws_raw.title = "Raw"
raw_header = ["subject", "from", "received", "preview"]
ws_raw.append(raw_header)
for r in raw_rows:
    ws_raw.append([r.get(h, "") for h in raw_header])

# Sheet 2: Parsed
ws_parsed = wb.create_sheet("Parsed")
# 基础列 + 动态字段列（按发现顺序）
parsed_header = ["_subject", "_from", "_received"] + all_keys
ws_parsed.append(parsed_header)

for r in parsed_rows:
    ws_parsed.append([r.get(h, "") for h in parsed_header])

# 美化列宽（简单自适应）
def autosize(ws):
    dims = defaultdict(int)
    for row in ws.iter_rows(values_only=True):
        for i, cell in enumerate(row, 1):
            s = "" if cell is None else str(cell)
            if len(s) > dims[i]:
                dims[i] = min(len(s), 80)  # 限制最大宽度
    for i, w in dims.items():
        ws.column_dimensions[get_column_letter(i)].width = max(10, w + 2)

autosize(ws_raw)
autosize(ws_parsed)

out_path = "outlook_parsed.xlsx"
wb.save(out_path)
print(f"已导出: {out_path}")
